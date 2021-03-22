import gspread
from time import time_ns
import re
from urllib.parse import urlparse
import pathlib
import time
import json
import sys
import openpyxl

from tracking import Group

header_aliases = {
	"Session ID": "session id",
	"World ID": "world id",
	"ticks played": "igt",
	"sprint one cm": "distance sprinted",
	"drop gold ingot": "gold dropped",
	"pickup blaze rod": "rods collected",
	"pickup flint": "flint collected",
	"pickup ender pearl": "pearls collected",
	"pickup gold ingot": "gold collected",
	"pickup iron ingot": "iron collected",
	"killed blaze": "blaze killed",
	"killed iron golem": "iron golem killed",
	"mined iron ore": "iron ore mined",
	"mined magma block": "magma block mined",
	"mined gravel": "gravel mined",
	
	"Wood": "getting wood",
	"Iron Pickaxe": "iron pickaxe",
	"Nether": "we need to go deeper",
	"Bastions": "those where the days",
	"Fortress": "a terrible fortress",
	"Nether Exit": "got ender eyes",
	"Stronghold": "eye spy",
	"End": "the end?",
	"IGT": "igt",
	"Gold Dropped": "gold dropped",
	"Blaze Rods": "rods collected",
	"Blazes": "blaze killed",
	"Dmg Taken": "damage taken",
	"Sprint Dist": "distance sprinted",
	"Flint": "flint collected",
	"Gravel": "gravel mined",
	"Prismarine": "mined prismarine",
	"Furnace": "furnace placed",
	"Iron Mined": "iron ore mined",
	"Hay Mined": "mined hay block",
	"IGs Killed": "iron golem killed",
	"Magma Block": "magma block mined",
}

local_path = pathlib.Path(__file__).parent.parent

def is_google_sheet(sheet):
	return urlparse(sheet).scheme != ""

def get_sheet_access(client, sheet):
	try:
		client.open_by_url(sheet)
		return True
	except:
		return False

def get_credentials(sheets, prompt):
	urls = [sheet for sheet in sheets if urlparse(sheet).scheme != ""]
	print("Credentials not found. please move credentials to exactly \"" + local_path.joinpath("credentials.json")
            .absolute().as_posix() + "\" and press enter, or type in the path that the credentials file is located at.")
	while True:
		if prompt:
			file = input("> ")
		else:
			path = local_path.joinpath("credentials.json").absolute()
			time.sleep(5)

		if file == "":
			path = local_path.joinpath("credentials.json").absolute()

		file = pathlib.Path(file)
		file = file.expanduser()

		if not file.exists():
			continue

		filename = file.as_posix()

		try:
			client = gspread.service_account(filename=filename)
		except:
			continue

		if all(get_sheet_access(client, sheet) for sheet in urls):
			return filename
		else:
			with file.open("r") as f:
				client_email = json.load(f)["client_email"]
				print("client couldnt access all target sheets. please make sure it is shared with " + client_email)

def get_client(filename):
	if filename == None:
		return None
	return gspread.service_account(filename=filename)

def create_sheet(uri, client):
	if is_google_sheet(uri):
		if client == None:
			print("ERROR: no client for google sheets")
			sys.exit(2)
		return GoogleSheet(uri, client)
	return SpreadSheet(uri)

class Sheet:
	columns = Group.get_names()
	column_count = len(columns)

	def __init__(self, sheet):
		self.sheet = sheet
		self.session_id = time_ns() // 1_000_000

		self.row_lines = []
		self.rows = []

		self.next_row = 0

	def stop(self):
		for row in range(len(self.rows)):
			self.update_row(row)

	def get_next_row(self):
		self.next_row += 1
		return self.next_row

	def get_active_row(self, folder_id):
		if len(self.rows) <= folder_id:
			self.rows += [ [None] * Sheet.column_count for i in range(len(self.row_lines) + 1 - folder_id)]
			self.row_lines += [ self.get_next_row() for i in range(len(self.row_lines) + 1 - folder_id)]
		return self.rows[folder_id]

	def update_row(self, folder_id, world_id, column_name, value):
		pass
	
	def push_row(self, folder_id):
		self.update_row(folder_id)
		self.rows[folder_id] = [None] * Sheet.column_count
		self.row_lines[folder_id] = self.get_next_row()

	def set_row_value(self, folder_id, column_name, value):
		try:
			self.get_active_row(folder_id)[Sheet.columns.index(column_name)] = value
		except:
			pass

	def update_values(self, folder_id, world_id, values):
		row = self.get_active_row(folder_id)
		last_world_id = row[2]
		if not last_world_id == world_id:
			if not last_world_id == None:
				self.push_row(folder_id)
			# set the meta data on this row
			self.set_row_value(folder_id, "session id", self.session_id)
			self.set_row_value(folder_id, "folder id", folder_id)
			self.set_row_value(folder_id, "world id", world_id)

		for value in values:
			self.set_row_value(folder_id, value.name, value.value)
		self.update_row(folder_id)

class SpreadSheet(Sheet):
	def __init__(self, filename):
		super().__init__(filename)
		self.workbook = openpyxl.load_workbook(filename)
		self.worksheet = self.workbook.get_sheet_by_name('Raw Data')
		self.next_row = self.worksheet.max_row + 1
	
	def update_row(self, folder_id):
		row = self.rows[folder_id]
		target_row = self.row_lines[folder_id]

		for i in range(len(row)):
			self.worksheet.cell(row=target_row, column=i + 1).value = row[i]
		
		self.workbook.save(self.sheet)

class GoogleSheet(Sheet):
	def __init__(self, url, client):
		super().__init__(client.open_by_url(url))
		self.worksheet = self.sheet.worksheet("Raw Data")
		self.next_row = len(self.worksheet.get("A1:A"))

	def update_row(self, folder_id):
		row = self.rows[folder_id]
		target_row = self.row_lines[folder_id]

		range_start = gspread.utils.rowcol_to_a1(target_row, 1)
		range_end = gspread.utils.rowcol_to_a1(target_row, len(row) + 1)
		try:
			self.worksheet.update(range_start + ":" + range_end, [row])
		except:
			self.worksheet.append_row(row)
