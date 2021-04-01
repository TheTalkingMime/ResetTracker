import gspread
import openpyxl

from urllib.parse import urlparse
import pathlib

import json
import sys

from threading import Lock, Thread

from time import time_ns
from time import sleep

from tracking import Group

header_aliases = {
	"Session ID": "session id",
	"World ID": "world id",
	"ticks played": "igt",
	"sprint one cm": "distance sprinted",
	"drop gold ingot": "gold dropped",
	"pickup blaze rod": "rods collected",
	"pickup flint": "flint collected",
	"pickup ender pearl": "pearls collected",
	"pickup gold ingot": "gold collected",
	"pickup iron ingot": "iron collected",
	"killed blaze": "blaze killed",
	"killed iron golem": "iron golem killed",
	"mined iron ore": "iron ore mined",
	"mined magma block": "magma block mined",
	"mined gravel": "gravel mined",
	
	"Wood": "getting wood",
	"Iron Pickaxe": "iron pickaxe",
	"Nether": "we need to go deeper",
	"Bastions": "those where the days",
	"Fortress": "a terrible fortress",
	"Nether Exit": "got ender eyes",
	"Stronghold": "eye spy",
	"End": "the end?",
	"IGT": "igt",
	"Gold Dropped": "gold dropped",
	"Blaze Rods": "rods collected",
	"Blazes": "blaze killed",
	"Dmg Taken": "damage taken",
	"Sprint Dist": "distance sprinted",
	"Flint": "flint collected",
	"Gravel": "gravel mined",
	"Prismarine": "mined prismarine",
	"Furnace": "furnace placed",
	"Iron Mined": "iron ore mined",
	"Hay Mined": "mined hay block",
	"IGs Killed": "iron golem killed",
	"Magma Block": "magma block mined",
}

local_path = pathlib.Path(__file__).parent.parent

def is_google_sheet(sheet):
	return urlparse(sheet).scheme != ""

def get_sheet_access(client, sheet):
	try:
		client.open_by_url(sheet)
		return True
	except:
		return False

def get_credentials(sheets, prompt):
	urls = [sheet for sheet in sheets if urlparse(sheet).scheme != ""]
	print("Credentials not found. please move credentials to exactly \"" + local_path.joinpath("credentials.json")
            .absolute().as_posix() + "\" and press enter, or type in the path that the credentials file is located at.")
	while True:
		if prompt:
			file = input("> ")
		else:
			path = local_path.joinpath("credentials.json").absolute()
			sleep(5)

		if file == "":
			path = local_path.joinpath("credentials.json").absolute()

		file = pathlib.Path(file)
		file = file.expanduser()

		if not file.exists():
			continue

		filename = file.as_posix()

		try:
			client = gspread.service_account(filename=filename)
		except:
			continue

		if all(get_sheet_access(client, sheet) for sheet in urls):
			return filename
		else:
			with file.open("r") as f:
				client_email = json.load(f)["client_email"]
				print("client couldnt access all target sheets. please make sure it is shared with " + client_email)

def get_client(filename):
	if filename == None:
		return None
	return gspread.service_account(filename=filename)

def create_sheet(uri, client):
	if is_google_sheet(uri):
		if client == None:
			print("ERROR: no client for google sheets")
			sys.exit(2)
		return GoogleSheet(uri, client)
	return SpreadSheet(uri)

class Sheet:
	columns = Group.get_names()
	column_count = len(columns)

	sheets = []

	@staticmethod
	def push_values(folder_id, world_id, values):
		for sheet in Sheet.sheets:
			thread = Thread(target = sheet.update_values, args = (folder_id, world_id, values))
			thread.start()

	@staticmethod
	def create_sheets(sheets, client):
		threads = [Thread(target=Sheet.create_sheet, args=(sheet, client)) for sheet in sheets]
		for thread in threads:
			thread.start()
		for thread in threads:
			thread.join()

	@staticmethod
	def create_sheet(uri, client):
		if is_google_sheet(uri):
			if client == None:
				print("ERROR: no client for google sheets")
				sys.exit(2)
			return GoogleSheet(uri, client)
		return SpreadSheet(uri)

	def __init__(self, sheet):
		Sheet.sheets.append(self)
		self.sheet = sheet
		self.session_id = time_ns() // 1_000_000

		self.row_lines = []
		self.rows = []

		self.next_row = 0

		self.orginize_columns()

		self.lock = Lock()

		self.queue = 0
		self.queue_lock = Lock()

		print("sheet added: " + sheet)

	def stop(self):
		for row in range(len(self.rows)):
			self.update_row(row)

	def get_next_row(self):
		self.next_row += 1
		return self.next_row

	def get_active_row(self, folder_id):
		if len(self.rows) <= folder_id:
			self.rows += [ [None] * Sheet.column_count for i in range(len(self.row_lines) + 1 - folder_id)]
			self.row_lines += [ self.get_next_row() for i in range(len(self.row_lines) + 1 - folder_id)]
		return self.rows[folder_id]

	# function template for updating a row on the sheet
	def update_row(self, folder_id, world_id, column_name, value):
		raise NotImplementedError

	def get_headers(self):
		raise NotImplementedError

	def get_range(self, start_x, start_y, end_x, end_y):
		raise NotImplementedError

	def set_range(self, start_x, start_y, end_x, end_y, data):
		raise NotImplementedError

	# function template for orginizing the rows on a sheet to match the headers we have
	def orginize_columns(self):
		headers = self.get_headers()
	
		current_headers = [ header if not header in header_aliases else header_aliases[header] for header in headers ]
		
		all_headers = list(dict.fromkeys(current_headers + Sheet.columns))

		current_data = self.get_data(len(all_headers))

		group_names = [group.name for group in Group.groups]

		groups = [[] for i in range(len(group_names))]
		no_group = []

		for index, header in enumerate(all_headers):
			if header in Sheet.columns:
				groups[group_names.index(Group.find_host_group(header))].append((header, index))
			else:
				no_group.append((header, index))

		for index, group in enumerate(groups):
			def get_group_order(e):
				g = Group.groups[index]
				for i, item in enumerate(g.get_items()):
					if e[0] == item.name:
						return i
				return 0

			group = group.sort(key=get_group_order)
		
		remapped_data = [item for group in groups for item in group] + no_group
		remapped_data = [[header, "-"] + current_data[index] for header, index in remapped_data ]

		self.set_data(remapped_data)
	
	def push_row(self, folder_id):
		self.update_row(folder_id)
		self.rows[folder_id] = [None] * Sheet.column_count
		self.row_lines[folder_id] = self.get_next_row()

	def set_row_value(self, folder_id, column_name, value):
		try:
			self.get_active_row(folder_id)[Sheet.columns.index(column_name)] = value
		except:
			pass

	def update_values(self, folder_id, world_id, values):
		with self.queue_lock:
			self.queue += 1

		# race condition protection
		with self.lock:
			row = self.get_active_row(folder_id)
			last_world_id = row[2]
			if not last_world_id == world_id:
				if not last_world_id == None:
					self.push_row(folder_id)
				# set the meta data on this row
				self.set_row_value(folder_id, "session id", self.session_id)
				self.set_row_value(folder_id, "folder id", folder_id)
				self.set_row_value(folder_id, "world id", world_id)

			for value in values:
				self.set_row_value(folder_id, value.name, value.value)
			self.update_row(folder_id)
			with self.queue_lock:
				self.queue -= 1

class SpreadSheet(Sheet):
	def __init__(self, filename):
		self.workbook = openpyxl.load_workbook(filename)
		self.worksheet = self.workbook.get_sheet_by_name('Raw Data')
		super().__init__(filename)
		self.next_row = self.worksheet.max_row + 1
	
	def update_row(self, folder_id):
		row = self.rows[folder_id]
		target_row = self.row_lines[folder_id]

		for column, value in enumerate(row):
			self.worksheet.cell(row=target_row, column=column + 1).value = value
		
		self.workbook.save(self.sheet)

	def get_headers(self):
		headers = list(self.worksheet.rows)
		if not len(headers) == 0:
			headers = [cell.value for cell in headers[0]]
		return headers

	def get_data(self, columns):
		data = [[cell.value if not cell.value == None else "" for cell in list(column)[2:]] for column in self.worksheet.iter_cols()]

		if len(data) == 0:
			row_count = 0
		else:
			row_count = len(data[0])

		data += [[""] * row_count for i in range(columns - len(data))]
		
		return data

	def set_data(self, data):
		for column, i in enumerate(data):
			for row, value in enumerate(i):
				self.worksheet.cell(row=row + 1, column=column + 1).value = value

		self.workbook.save(self.sheet)

class GoogleSheet(Sheet):
	def __init__(self, url, client):
		sheet = client.open_by_url(url)
		self.worksheet = sheet.worksheet("Raw Data")
		super().__init__(url)
		self.next_row = len(self.worksheet.get("A1:A"))

	def update_row(self, folder_id):
		if not self.queue == 1:
			return

		row = self.rows[folder_id]
		target_row = self.row_lines[folder_id]

		range_start = gspread.utils.rowcol_to_a1(target_row, 1)
		range_end = gspread.utils.rowcol_to_a1(target_row, len(row) + 1)

		success = False
		while not success:
			try:
				self.worksheet.update(range_start + ":" + range_end, [row])
				success = True
			except gspread.exceptions.GSpreadException:
				self.worksheet.append_row(row)
				success = True
			except gspread.exceptions.APIError:
				print("api rate limit error, trying again in 15 seconds")
				sleep(15)

	def get_range_a1(self, start_x, start_y, end_x, end_y):
		if (start_x == None or start_x < 1) and (start_y == None or start_y < 1):
			raise RuntimeError
		elif start_x == None:
			a1_start = gspread.utils.rowcol_to_a1(start_y, 1)[1:]
		elif start_y == None:
			a1_start = gspread.utils.rowcol_to_a1(1, start_x)[:-1]
		else:
			a1_start = gspread.utils.rowcol_to_a1(start_y, start_x)

		if (end_x == None or end_x < 1) and (end_y == None or end_y < 1):
			raise RuntimeError
		elif end_x == None:
			a1_end = gspread.utils.rowcol_to_a1(end_y, 1)[1:]
		elif end_y == None:
			a1_end = gspread.utils.rowcol_to_a1(1, end_x)[:-1]
		else:
			a1_end = gspread.utils.rowcol_to_a1(end_y, end_x)

		return a1_start + ":" + a1_end

	def expand_sheet(self, width, height):
		if width > self.worksheet.row_count:
			missing_row = height - self.worksheet.row_count
			self.worksheet.add_rows(missing_row)

		if height > self.worksheet.col_count:
			missing_col = width - self.worksheet.col_count
			self.worksheet.add_cols(missing_col)

	def get_headers(self):
		headers = self.worksheet.get("A1:" + gspread.utils.rowcol_to_a1(1, self.worksheet.col_count))
		try:
			return headers[0]
		except:
			return headers

	def get_data(self, columns):
		data = self.worksheet.batch_get([self.get_range_a1(i, 3, i, None) for i in range(1, self.worksheet.col_count + 1)])
		data = [[row[0] if not len(row) == 0 else "" for row in column] + [""] * (self.worksheet.row_count - len(column) - 2) for column in data]

		data += [[""] * self.worksheet.row_count for i in range(columns - self.worksheet.col_count)]

		def parse_value(value):
			try:
				return float(value)
			except:
				return value

		data = [[ parse_value(cell) for cell in column ] for column in data]

		return data

	def set_data(self, data):
		self.expand_sheet(len(data), len(data[0]))

		self.worksheet.batch_update([{
			"range": gspread.utils.rowcol_to_a1(1, index + 1) + ":" + gspread.utils.rowcol_to_a1(1, index + 1)[:-1],
			"values": [ [ row ] for row in column ]
		} for index, column in enumerate(data)], value_input_option="RAW")
